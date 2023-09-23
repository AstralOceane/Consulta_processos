from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl
import hashlib
import smtplib
from email.mime.text import MIMEText

# Função para calcular o hash de uma string
def calcular_hash(string):
    return hashlib.sha256(string.encode()).hexdigest()

# Função para enviar e-mail
def enviar_email(mensagem):
    de = 'seu_email@gmail.com'
    senha = 'sua_senha'
    para = 'destinatario@gmail.com'

    msg = MIMEText(mensagem)
    msg['From'] = de
    msg['To'] = para
    msg['Subject'] = 'Alerta de alterações na página'

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(de, senha)
    server.sendmail(de, para, msg.as_string())
    server.quit()

# Inicializar o WebDriver do Chrome
driver = webdriver.Chrome()
driver.get('https://pje-consulta-publica.tjmg.jus.br/')
sleep(15)
hash_inicial = None

# Digitar o número da OAB
numero_oab = 133864
campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

# Selecionar o estado
dropdown_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
opcoes_estados.select_by_visible_text('SP')

# Clicar em pesquisar
botao_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
botao_pesquisar.click()
sleep(10)

# Entrar em cada um dos processos
processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")

try:
    workbook = openpyxl.load_workbook('dados.xlsx')
except FileNotFoundError:
    workbook = openpyxl.Workbook()

for processo in processos:
    processo.click()
    sleep(10)
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1920, 1080)

    numero_processo_element = driver.find_element(By.XPATH, "//div[@class='col-sm-12 ']")
    data_distribuicao_element = driver.find_element(By.XPATH, "//div[@class='value col-sm-12 ']")

    numero_processo = numero_processo_element.text
    data_distribuicao = data_distribuicao_element.text

    movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = [movimentacao.text for movimentacao in movimentacoes]

    try:
        pagina_processo = workbook[numero_processo]
    except KeyError:
        pagina_processo = workbook.create_sheet(numero_processo)

    pagina_processo['A1'] = "Número Processo"
    pagina_processo['B1'] = "Data de Distribuição"
    pagina_processo['C1'] = "Movimentações"
    pagina_processo['A2'] = numero_processo
    pagina_processo['B2'] = data_distribuicao

    for index, movimentacao in enumerate(lista_movimentacoes, start=3):
        pagina_processo.cell(row=index, column=3, value=movimentacao)

    workbook.save('dados.xlsx')

    # Verifica se houve alterações na página
    hash_atual = calcular_hash(driver.page_source)
    if hash_inicial is not None and hash_atual != hash_inicial:
        # Se houve alterações, enviar um e-mail
        enviar_email("Alterações detectadas na página!")

    hash_inicial = hash_atual

    driver.close()
    sleep(5)
    driver.switch_to.window(driver.window_handles[0])

# Fechar o WebDriver
driver.quit()
